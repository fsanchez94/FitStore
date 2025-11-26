#!/usr/bin/env python
"""Script to import products from Book1.xlsx into the database"""

import os
import sys
import django
import openpyxl
from decimal import Decimal

# Setup Django
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
os.environ.setdefault('DJANGO_SETTINGS_MODULE', 'fitstore_backend.settings')
django.setup()

from supplements.models import Product

def import_products():
    # Path to the Excel file
    excel_file = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'Book1.xlsx')

    if not os.path.exists(excel_file):
        print(f"Error: File not found at {excel_file}")
        return

    # Load the workbook
    wb = openpyxl.load_workbook(excel_file)
    ws = wb.active

    print(f"Reading from sheet: {ws.title}")
    print(f"Total rows: {ws.max_row}")
    print()

    # Get headers from first row
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)

    print(f"Headers: {headers}")
    print()

    # Process each row (skip header)
    products_created = 0
    products_updated = 0

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):  # Skip empty rows
            continue

        # Create a dictionary from the row
        row_data = dict(zip(headers, row))

        print(f"Row {row_idx}: {row_data}")

        # Map Excel columns to model fields
        # Adjust these mappings based on your Excel file structure
        product_data = {
            'product_name': row_data.get('Product Name') or row_data.get('product_name') or '',
            'product_type': row_data.get('Product Type') or row_data.get('product_type') or '',
            'url': row_data.get('URL') or row_data.get('url') or '',
            'amount_per_serving': str(row_data.get('Amount Per Serving') or row_data.get('Amount per serving') or row_data.get('amount_per_serving') or ''),
            'serving_size': str(row_data.get('Serving Size') or row_data.get('serving_size') or ''),
            'description': row_data.get('Description') or row_data.get('description') or '',
            'unit': row_data.get('Unit') or row_data.get('unit') or 'unit',
        }

        # Handle numeric fields
        units_per_container = row_data.get('Units per Container') or row_data.get('Units Per Container') or row_data.get('units_per_container')
        if units_per_container is not None:
            try:
                product_data['units_per_container'] = Decimal(str(units_per_container))
            except:
                product_data['units_per_container'] = None
        else:
            product_data['units_per_container'] = None

        weight_g = row_data.get('Weight G') or row_data.get('Weight(g)') or row_data.get('weight_g') or row_data.get('Weight (g)')
        if weight_g is not None:
            try:
                product_data['weight_g'] = Decimal(str(weight_g))
            except:
                product_data['weight_g'] = None
        else:
            product_data['weight_g'] = None

        current_stock = row_data.get('Current Stock') or row_data.get('current_stock')
        if current_stock:
            try:
                product_data['current_stock'] = Decimal(str(current_stock))
            except:
                product_data['current_stock'] = Decimal('0')
        else:
            product_data['current_stock'] = Decimal('0')

        min_stock = row_data.get('Min Stock Level') or row_data.get('min_stock_level')
        if min_stock:
            try:
                product_data['min_stock_level'] = Decimal(str(min_stock))
            except:
                product_data['min_stock_level'] = Decimal('0')
        else:
            product_data['min_stock_level'] = Decimal('0')

        # Skip if product name is empty
        if not product_data['product_name']:
            print(f"  Skipping row {row_idx}: No product name")
            continue

        # Check if product already exists
        existing_product = Product.objects.filter(product_name=product_data['product_name']).first()

        if existing_product:
            # Update existing product
            for key, value in product_data.items():
                setattr(existing_product, key, value)
            existing_product.save()
            products_updated += 1
            print(f"  Updated: {product_data['product_name']}")
        else:
            # Create new product
            product = Product.objects.create(**product_data)
            products_created += 1
            print(f"  Created: {product_data['product_name']}")

        print()

    print(f"\nImport complete!")
    print(f"Products created: {products_created}")
    print(f"Products updated: {products_updated}")
    print(f"Total products in database: {Product.objects.count()}")

if __name__ == '__main__':
    import_products()
